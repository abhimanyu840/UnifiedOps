from __future__ import annotations

import asyncio
import csv
import io
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

from .influx_pool import InfluxPool, InfluxQueryError
from .alert_monitor import severity_from_body
from .dashboard import (
    scoped_buckets,
    bucket_key,
    range_clause,
    strip_syslog_header,
    bucket_severity,
    _MEASUREMENT_FILTER,
    _COUNT_FIELD_FILTER,
)

from .device_inventory import _load_inventory_json, _DEFAULT_INVENTORY_PATH

log = logging.getLogger("unifiedops.reports")

HEALTH_CHECK_CATEGORIES = [
    "ntp_alert", "fips_alert", "wwn_alert", 
    "airflow_alert", "battery_alert", "fan_missing", "fan_failure", "blade_fault", 
    "temperature_alarm", "voltage_alert", "power_failure", "optic_alert", 
    "firmware_alert", "flow_vision", "config_change", "ha_alert", "port_fault", 
    "chassis_alert", "fru_event", "env_warning"
]

import re
_HC_PATTERNS = [
    ("ntp_alert", re.compile(r"\bntp\b|\bsntp\b|time\s*sync|clock\s*(drift|skew|sync)", re.I)),
    ("fips_alert", re.compile(r"\bfips\b", re.I)),
    ("wwn_alert", re.compile(r"\bwwn\b|world[\s_-]*wide[\s_-]*name|wwn conflict", re.I)),
    ("airflow_alert", re.compile(r"air[\s_-]*flow|psu-?fan", re.I)),
    ("battery_alert", re.compile(r"\bbattery\b", re.I)),
    ("fan_missing", re.compile(r"missing\s*(fan|blower)|fan\s*missing", re.I)),
    ("fan_failure", re.compile(r"\b(fan|blower)(s)?\b.*\b(fail(ed|s)?|fault(ed|y)?|stop(ped)?)\b|fan failure", re.I)),
    ("temperature_alarm", re.compile(r"\b(temp(erature)?|overheat|thermal|cool(ing)?)\b", re.I)),
    ("voltage_alert", re.compile(r"\bvoltage\b", re.I)),
    ("power_failure", re.compile(r"\bpower\s*(fail|loss|down|fault)|AC\s*(fail|loss)", re.I)),
    ("optic_alert", re.compile(r"\bsfp\b|\bqsfp\b|\boptic", re.I)),
    ("firmware_alert", re.compile(r"\b(firmware|fpga|fos)\b", re.I)),
    ("flow_vision", re.compile(r"flow\s*vision", re.I)),
    ("config_change", re.compile(r"\bconfig(uration)?\b.*\b(change|update|modif)\b|config changed", re.I)),
    ("ha_alert", re.compile(r"\bfailover\b|ha failover|ha\s*(reboot|sync)", re.I)),
    ("port_fault", re.compile(r"\bport\b.*\b(fail|fault|down)\b", re.I)),
    ("blade_fault", re.compile(r"\bblade\b.*\b(fail|fault|offline)\b", re.I)),
    ("chassis_alert", re.compile(r"\bchassis\b.*\b(disable|fault|error)\b", re.I)),
    ("fru_event", re.compile(r"\bfru\b.*\b(fault|error|fail)\b", re.I)),
    ("env_warning", re.compile(r"environment(al)?\s*warning", re.I)),
]

def _reclassify_event(text: str, default_cat: str = "sannav_event") -> str:
    for cat, pattern in _HC_PATTERNS:
        if pattern.search(text):
            return cat
    return default_cat

def _flux_report(bucket: str, range_key: str, limit: int = 50000) -> str:
    """Pull raw messages up to a large limit for export."""
    return (
        f'from(bucket: "{bucket}")\n'
        f'  |> range({range_clause(range_key)})\n'
        f'  |> filter(fn: (r) => {_MEASUREMENT_FILTER})\n'
        f'  |> filter(fn: (r) => {_COUNT_FIELD_FILTER})\n'
        f'  |> sort(columns: ["_time"], desc: true)\n'
        f'  |> limit(n: {limit})\n'
    )

class ReportService:
    def __init__(self, pool: InfluxPool) -> None:
        self._pool = pool

    async def get_multi_format_report(
        self,
        range_key: str,
        sites: Optional[List[str]] = None,
        vendors: Optional[List[str]] = None,
        limit: int = 50000,
        fmt: str = "csv",
        report_type: str = "hardware",
    ) -> tuple[bytes, str, str]:
        """Fetch alerts across requested sites/vendors and return as formatted bytes, along with media_type and extension."""
        import io
        import csv
        import zipfile
        from collections import defaultdict
        import openpyxl
        from fpdf import FPDF

        buckets = [dict(b) for b in scoped_buckets(sites, vendors)]
        buckets = [dict(b) for b in scoped_buckets(sites, vendors)]
        target_sites = []
        target_vendors = []
        
        if report_type == "health_check":
            import os
            for b in buckets:
                target_sites.append(b["site"].upper())
                target_vendors.append(b["vendor"].lower())
                
                if b["vendor"] == "brocade":
                    site_upper = b["site"].upper()
                    b["bucket"] = os.environ.get(f"HITRACK_INFLUX_BROCADE_{site_upper}_REPORT_BUCKET", f"unified-ops-bucket-health-check-report-{b['site'].lower()}")
                    report_url = os.environ.get(f"HITRACK_INFLUX_BROCADE_{site_upper}_REPORT_URL")
                    report_token = os.environ.get(f"HITRACK_INFLUX_BROCADE_{site_upper}_REPORT_TOKEN")
                    if report_url:
                        b["url"] = report_url
                    if report_token:
                        b["token"] = report_token
                    
                    # Register the report bucket in the pool dynamically
                    report_key = f"report:{b['site']}:{b['vendor']}"
                    self._pool.register(report_key, url=b.get("url"), token=b.get("token"), org=b["org"])

            target_sites = list(set(target_sites))
            target_vendors = list(set(target_vendors))

        per_bucket = await asyncio.gather(*[
            self._safe_query(
                f"report:{b['site']}:{b['vendor']}" if (report_type == "health_check" and b["vendor"] == "brocade") else bucket_key(b),
                _flux_report(b["bucket"], range_key, limit)
            )
            for b in buckets
        ])

        out: List[Dict[str, Any]] = []
        for cfg, rows in zip(buckets, per_bucket):
            vendor = cfg["vendor"].lower()
            for r in rows:
                ts = r.get("_time")
                if not ts:
                    continue
                try:
                    dt = datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
                except Exception:
                    continue
                
                local_time = dt.astimezone().strftime("%Y-%m-%d %H:%M:%S")
                
                storage = (
                    r.get("array_name")
                    or r.get("switch_name")
                    or r.get("hostname")
                    or r.get("source_ip", "-")
                )
                if storage in ("", "unknown", "-"):
                    storage = r.get("source_ip", "-")
                
                raw_body = str(r.get("_value") or "")
                event_text = strip_syslog_header(raw_body) or raw_body
                
                raw_sev = r.get("severity")
                severity = bucket_severity(raw_sev) if raw_sev else "informational"
                if not raw_sev or severity == "informational":
                    body_sev = severity_from_body(raw_body)
                    if body_sev is not None:
                        severity = body_sev
                
                trap_cat = (r.get("trap_category") or "other")
                if report_type == "health_check":
                    trap_cat = _reclassify_event(event_text, trap_cat)
                elif trap_cat == "sannav_event" or vendor == "sannav":
                    reclass = _reclassify_event(event_text, "sannav_event")
                    if reclass != "sannav_event":
                        trap_cat = reclass

                out.append({
                    "Timestamp": ts,
                    "Local Time": local_time,
                    "Severity": severity.capitalize(),
                    "Storage/Switch": storage,
                    "Source IP": r.get("source_ip") or "-",
                    "Event Details": event_text,
                    "Raw Syslog": raw_body,
                    "Category": trap_cat.capitalize(),
                    "Location": cfg["site"],
                    "Vendor": vendor.capitalize(),
                })
        
        # Sort by actual timestamp descending
        out.sort(key=lambda a: a["Timestamp"], reverse=True)
        
        print(f"DEBUG OUT LENGTH IN BACKEND: {len(out)}", flush=True)

        fieldnames = [
            "Local Time", "Location", "Vendor", "Storage/Switch",
            "Severity", "Category", "Source IP", "Event Details"
        ]
        
        report_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if report_type == "health_check":
            excel_bytes = self._generate_health_check_excel(out, report_time_str, target_sites, target_vendors)
            return excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"

        if not out:
            if fmt == "xlsx":
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "No Data"
                ws.append([f"Report Generated At: {report_time_str}"])
                ws.append([])
                ws.append(["No data found for the selected criteria."])
                buf = io.BytesIO()
                wb.save(buf)
                return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
            elif fmt == "pdf":
                buf = io.BytesIO()
                with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    zf.writestr("No_Data.txt", f"Report Generated At: {report_time_str}\n\nNo data found for the selected criteria.".encode('utf-8'))
                return buf.getvalue(), "application/zip", "zip"
            else:
                return f"Report Generated At: {report_time_str}\n\nNo data found for the selected criteria.\n".encode('utf-8'), "text/csv", "csv"

        # Group data by Vendor_Site
        groups = defaultdict(list)
        for row in out:
            tab_name = f"{row['Vendor']}_{row['Location']}"
            groups[tab_name].append(row)

        if fmt == "xlsx":
            wb = openpyxl.Workbook()
            # Remove default active sheet if we are going to create our own
            if len(wb.sheetnames) > 0:
                del wb[wb.sheetnames[0]]

            for tab_name, rows in groups.items():
                ws = wb.create_sheet(title=tab_name[:31])
                ws.append([f"Report Generated At: {report_time_str}"])
                ws.append([])
                ws.append(fieldnames)
                for r in rows:
                    ws.append([str(r[f]) for f in fieldnames])

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"

        elif fmt == "pdf":
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for tab_name, rows in groups.items():
                    pdf = FPDF(orientation="landscape")
                    pdf.set_auto_page_break(auto=True, margin=15)
                    pdf.add_page()
                    
                    pdf.set_font("helvetica", "B", 12)
                    pdf.cell(0, 10, f"UnifiedOps Alerts: {tab_name}", new_x="LMARGIN", new_y="NEXT", align="C")
                    
                    pdf.set_font("helvetica", "I", 10)
                    pdf.cell(0, 10, f"Report Generated At: {report_time_str}", new_x="LMARGIN", new_y="NEXT", align="L")
                    pdf.ln(2)
                    
                    # Define headers to display in PDF
                    pdf_headers = ["Local Time", "Severity", "Storage/Switch", "Source IP", "Category", "Event Details"]
                    
                    with pdf.table(col_widths=(30, 20, 40, 25, 25, 137), text_align="LEFT") as table:
                        header_row = table.row()
                        pdf.set_font("helvetica", "B", 9)
                        for header in pdf_headers:
                            header_row.cell(header)
                        
                        pdf.set_font("helvetica", "", 8)
                        for r in rows:
                            data_row = table.row()
                            data_row.cell(str(r["Local Time"]))
                            data_row.cell(str(r["Severity"]))
                            data_row.cell(str(r["Storage/Switch"]))
                            data_row.cell(str(r["Source IP"]))
                            data_row.cell(str(r["Category"]))
                            data_row.cell(str(r["Event Details"]))
                        
                    pdf_bytes = pdf.output()
                    zf.writestr(f"{tab_name}.pdf", pdf_bytes)

            return buf.getvalue(), "application/zip", "zip"

        else:
            # Default to CSV
            output = io.StringIO()
            output.write(f"Report Generated At: {report_time_str}\n\n")
            writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            for row in out:
                writer.writerow(row)
            return output.getvalue().encode('utf-8'), "text/csv", "csv"

    async def _safe_query(self, key: str, flux: str) -> List[Dict[str, Any]]:
        try:
            return await self._pool.query(key, flux)
        except InfluxQueryError as exc:
            log.warning("reports %s query failed: %s", key, exc.reason)
            raise
        except Exception as exc:
            log.warning("reports %s query crash: %s", key, exc)
            raise

    def _generate_health_check_excel(self, out: List[Dict[str, Any]], report_time_str: str, target_sites: List[str], target_vendors: List[str]) -> bytes:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from collections import defaultdict
        
        inv = _load_inventory_json(_DEFAULT_INVENTORY_PATH)
        
        red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        bold_font = Font(bold=True)
        white_bold_font = Font(color="FFFFFF", bold=True)
        green_bold_font = Font(color="00B050", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        wb = openpyxl.Workbook()
        if len(wb.sheetnames) > 0:
            del wb[wb.sheetnames[0]]
            
        alerts_by_site_vendor = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list))))
        for r in out:
            v = r["Vendor"].lower()
            s = r["Location"].upper()
            sw = r["Storage/Switch"]
            cat = r["Category"].lower()
            if v == "sannav":
                v = "brocade"
            alerts_by_site_vendor[s][v][sw][cat].append(r)
            
        for site in sorted(target_sites):
            for vendor in sorted(target_vendors):
                vendor_key = "brocade" if vendor == "sannav" else vendor
                vendor_inv = inv.get(vendor_key, {})
                site_devices = list(vendor_inv.get(site, []))
                seen_devices = set(site_devices)
                
                for sw in alerts_by_site_vendor[site][vendor_key].keys():
                    if sw not in seen_devices:
                        site_devices.append(sw)
                        seen_devices.add(sw)
                        
                if not site_devices:
                    continue
                    
                sheet_title = f"{site}_{vendor_key.capitalize()}_Summary"
                ws = wb.create_sheet(title=sheet_title[:31])
                ws.append([f"Health Check Report: {site} - {vendor_key.capitalize()}"])
                ws.append([f"Generated At: {report_time_str}"])
                ws.append([])
                
                headers = ["Device Name"] + HEALTH_CHECK_CATEGORIES
                ws.append(headers)
                for cell in ws[4]:
                    cell.font = bold_font
                    
                for sw in sorted(site_devices):
                    row = [sw]
                    for cat in HEALTH_CHECK_CATEGORIES:
                        count = len(alerts_by_site_vendor[site][vendor_key][sw][cat])
                        if count == 0:
                            row.append("✓")
                        else:
                            row.append(f"{count} Alerts")
                    ws.append(row)
                    
                    current_row = ws.max_row
                    for col_idx, cat in enumerate(HEALTH_CHECK_CATEGORIES, start=2):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.alignment = center_align
                        if cell.value == "✓":
                            cell.font = green_bold_font
                        else:
                            cell.fill = red_fill
                            cell.font = white_bold_font
                            
        drill_headers = ["Local Time", "Severity", "Category", "Event Details", "Raw Syslog"]
        
        for site in sorted(target_sites):
            for vendor in sorted(target_vendors):
                vendor_key = "brocade" if vendor == "sannav" else vendor
                for sw, cats in alerts_by_site_vendor[site][vendor_key].items():
                    all_sw_alerts = []
                    for cat_alerts in cats.values():
                        all_sw_alerts.extend(cat_alerts)
                        
                    if not all_sw_alerts:
                        continue
                        
                    all_sw_alerts.sort(key=lambda x: x["Timestamp"], reverse=True)
                    
                    safe_sw = sw.replace("/", "_").replace("\\", "_").replace("[", "_").replace("]", "_").replace("*", "_").replace("?", "_").replace(":", "_")
                    sheet_title = f"{safe_sw}"[:31]
                    
                    base_title = sheet_title
                    counter = 1
                    while sheet_title in wb.sheetnames:
                        suffix = f"_{counter}"
                        sheet_title = base_title[:31 - len(suffix)] + suffix
                        counter += 1
                    
                    ws = wb.create_sheet(title=sheet_title)
                    ws.append([f"Alert Details for {sw} ({site} - {vendor_key.capitalize()})"])
                    ws.append([])
                    ws.append(drill_headers)
                    for cell in ws[3]:
                        cell.font = bold_font
                        
                    for a in all_sw_alerts:
                        ws.append([
                            a["Local Time"],
                            a["Severity"],
                            a["Category"],
                            a["Event Details"],
                            a["Raw Syslog"]
                        ])
                        
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet(title="No Data")
            ws.append([f"Health Check Report Generated At: {report_time_str}"])
            ws.append([])
            ws.append(["No devices configured or found for the selected criteria."])
            
        import io
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
